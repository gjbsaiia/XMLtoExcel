#!python3
import os
import sys
import openpyxl
import xml.etree.ElementTree as ET

# data file name
wkbk = "picnicData.xlsx"
# music sheet name
mu = "Music"
# food sheet name
fu = "Food"
# activities sheet name
act = "Activities"

# folder that houses all xmls that need to be processed
xml_fold = "raw_data/"
# all music XML
mu_xml = []
# attributes belonging to music
mu_att = ["Song"]
# all food XML
fu_xml = []
# attributes belonging to food
fu_att = ["HamB", "VegB", "HDog", "VDog", "Ketchup", "Mustard", "Mayo", "Relish", "Lettuce", "Tomato", "Onion", "Pickles", "Cheese"]
# all activity XML
act_xml = []
# attributes belonging to activity
act_att = ["Name", "Member2", "Tug_Y_N", "Home_Y_N", "Pin_Y_N", "Rock", "HipHop", "Alternative", "RnB", "EDM", "Pop"]

# Maps XML tag to location on Excel Sheet
# Convention: XML = ["Column", LastIndex, IndexShouldBeUpdated]
indexMap = {
    "Name": ["A",2, True],
    "Member2": ["B",2, True],
    "Tug_Y_N": ["C",2, True],
    "Home_Y_N": ["D",2, True],
    "Pin_Y_N": ["E", 2, True],
    "Rock": ["L", 2, False],
    "HipHop": ["M", 2, False],
    "Alternative": ["N", 2, False],
    "RnB": ["O", 2, False],
    "EDM": ["P", 2, False],
    "Pop": ["Q", 2, False],
    "HamB": ["A", 2, False],
    "VegB": ["B", 2, False],
    "HDog": ["C", 2, False],
    "VDog": ["D", 2, False],
    "Ketchup": ["E", 2, False],
    "Mustard": ["F", 2, False],
    "Mayo": ["G", 2, False],
    "Relish": ["H", 2, False],
    "Lettuce": ["I", 2, False],
    "Tomato": ["J", 2, False],
    "Onion": ["K", 2, False],
    "Pickles": ["L", 2, False],
    "Cheese": ["M", 2, False],
    "Song": ["A", 2, True]
}

def main():
    newData()
    if mu_xml:
        updateIndexes(mu, mu_att)
        for each in mu_xml:
            values = stripMuXml(each, mu_att)
            writeExcel(mu, values, mu_att)
    if act_xml:
        updateIndexes(act, act_att)
        for each in act_xml:
            values = stripActXml(each, act_att)
            writeExcel(act, values, act_att)
    if fu_xml:
        for each in fu_xml:
            values = stripFuXml(each, fu_att)
            writeFoodExcel(fu, values, fu_att)
    wipeData()
    
# update indexes in IndexMap
def updateIndexes(name, att):
    for each in att:
        if indexMap[each][2]:
            updateEntry = { each : [indexMap[each][0], getMaxRow(name, indexMap[each][0]), indexMap[each][2]]}
            indexMap.update(updateEntry)

# checks for and organizes new data sets
def newData():
    global mu_xml, fu_xml, act_xml, xml_fold
    resp = os.listdir(xml_fold)
    for each in resp:
        if "Song" in each:
            mu_xml.append(xml_fold+each)
        if "food" in each:
            fu_xml.append(xml_fold+each)
        if "activities" in each:
            act_xml.append(xml_fold+each)

def wipeData():
    global xml_fold
    resp = os.listdir(xml_fold)
    for each in resp:
        os.remove(xml_fold+each)

# helper to return Max row cleanly
def getMaxRow(name, col):
    global wkbk
    wb = openpyxl.load_workbook(wkbk)
    sheet = wb.get_sheet_by_name(name)
    i = 2
    while(sheet[col+str(i)].value != None):
        i += 1
    return i

# wrapper method to handle XML extraction and writing data into excel
def writeExcel(name, values, att):
    global wkbk
    wb = openpyxl.load_workbook(wkbk)
    sheet = wb.get_sheet_by_name(name)
    i = 0
    for each in values:
        if(i >= len(att)):
            break
        if(indexMap[att[i]][2]):
            sheet[indexMap[att[i]][0]+str(indexMap[att[i]][1])] = each
            update = { att[i]: [indexMap[att[i]][0], indexMap[att[i]][1]+1, True]}
            indexMap.update(update)
        else:
            sheet[indexMap[att[i]][0]+str(indexMap[att[i]][1])] = int(each) + int(sheet[indexMap[att[i]][0]+str(indexMap[att[i]][1])].value)
        if( i < len(att)-1):
            i+=1
    wb.save("picnicData.xlsx")

def writeFoodExcel(name, values, att):
    global wkbk
    wb = openpyxl.load_workbook(wkbk)
    sheet = wb.get_sheet_by_name(name)
    i = 0
    for each in values:
        if i == 0 or i == 2:
            if(i == 0):
                if(int(each) < 0):
                    sheet[indexMap["VegB"][0]+str(indexMap["VegB"][1])] = abs(int(each)) + sheet[indexMap["VegB"][0]+str(indexMap["VegB"][1])].value
                else:
                    sheet[indexMap["HamB"][0]+str(indexMap["HamB"][1])] = abs(int(each)) + sheet[indexMap["HamB"][0]+str(indexMap["HamB"][1])].value
                if( i < len(att)-1):
                    i+=1
            else:
                if(int(each) < 0):
                    sheet[indexMap["VDog"][0]+str(indexMap["VDog"][1])] = abs(int(each)) + sheet[indexMap["VDog"][0]+str(indexMap["VDog"][1])].value
                else:
                    sheet[indexMap["HDog"][0]+str(indexMap["HDog"][1])] = abs(int(each)) + sheet[indexMap["HDog"][0]+str(indexMap["HDog"][1])].value
                if( i < len(att)-1):
                    i+=1
        else:
            sheet[indexMap[att[i]][0]+str(indexMap[att[i]][1])] = int(each) + int(sheet[indexMap[att[i]][0]+str(indexMap[att[i]][1])].value)
        if( i < len(att)-1):
            i+=1
    wb.save("picnicData.xlsx")

def stripActXml(xml, att):
    values = []
    tree = ET.parse(xml)
    root = tree.getroot()
    xmlstr = ET.tostring(root, encoding='utf8', method='xml')
    xmlLis = xmlstr.split("\n")
    j = 0
    for each in xmlLis:
        if(">" in each and "<" in each):
            splitted = each.split(">")
            if att[j] in splitted[0]:
                if splitted[1]:
                    if(splitted[1].split("<")[0] == "true"):
                        values.append("1")
                    else:
                        values.append(splitted[1].split("<")[0])
                else:
                    values.append("")
                if( j < len(att)-1):
                    j+=1
    return values

def stripMuXml(xml, att):
    values = []
    tree = ET.parse(xml)
    root = tree.getroot()
    xmlstr = ET.tostring(root, encoding='utf8', method='xml')
    xmlLis = xmlstr.split("\n")
    j = 0
    for each in xmlLis:
        if(">" in each and "<" in each):
            splitted = each.split(">")
            if splitted[1] and "field2" in splitted[1]:
                if "," in splitted[1].split("<")[0]:
                    splittted = splitted[1].split("<")[0].split(",")
                    for each in splittted:
                        values.append(each)
                else:
                    values.append(splitted[1].split("<")[0])
                break
    return values

def stripFuXml(xml, att):
    values = []
    tree = ET.parse(xml)
    root = tree.getroot()
    xmlstr = ET.tostring(root, encoding='utf8', method='xml')
    xmlLis = xmlstr.split("\n")
    j = 0
    next = False
    for each in xmlLis:
        v = 1
        if(">" in each and "<" in each):
            splitted = each.split(">")
            if(next):
                values.append(int(splitted[1].split("<")[0]) * v)
                next = False
                if( j < len(att)-1):
                    j+=1
            if ":Burger" in splitted[0] or ":Dogs" in splitted[0]:
                if(splitted[1].split("<")[0] == "true"):
                    v = -1
                else:
                    v = 1
                next = True
                if( j < len(att)-1):
                    j+=1
            if att[j] in splitted[0]:
                if splitted[1]:
                    if(splitted[1].split("<")[0] == "true"):
                        values.append("1")
                    else:
                        values.append(splitted[1].split("<")[0])
                else:
                    values.append("0")
                if( j < len(att)-1):
                    j+=1
    return values




if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        print('Interrupted \_[*.*]_/\n')
        try:
            sys.exit(0)
        except SystemExit:
            os.exit(0)